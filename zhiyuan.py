#下面这些变量需要您根据自己的具体情况选择 
biaotou=['院校代码','院校名称','专业代码','专业名称','专业备注','计划数','类型','性质','选科','省份','城市','学费','去年提档分数','去年提档位次'] 
#在哪里搜索多个表格 
#E:\My_files\Writer\gaokao\
filelocation="E:\\Tools\\GaoKao_files\\DATA2023\\"
#当前文件夹下搜索的文件名后缀 
fileform="xlsx"
#将合并后的表格存放到的位置 
filedestination="E:\\Tools\\GaoKao_files\\DATA2023\\"
#合并后的表格命名为file 
file="2023_pool.xlsx"


#首先查找默认文件夹下有多少文档需要整合 
import glob 
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook 

filearray=[] 
for filename in glob.glob(filelocation+"*."+fileform): 
	filearray.append(filename)
#以上是从pythonscripts文件夹下读取所有excel表格，并将所有的名字存储到列表filearray 
print("在默认文件夹下有%d个文档哦"%len(filearray)) 

#0 - E:\Tools\GaoKao_files\DATA2023\2022一分一档.xlsx
#1 - E:\Tools\GaoKao_files\DATA2023\2022历史提档分数.xlsx
#2 - E:\Tools\GaoKao_files\DATA2023\2022物理提档分数.xlsx
#4 - E:\Tools\GaoKao_files\DATA2023\2023年历史招生计划.xlsx
#5 - E:\Tools\GaoKao_files\DATA2023\2023年物理招生计划.xlsx
#实现读写数据 
  
#下面是将所有文件读数据到三维列表cell[][][]中（不包含表头） 


ref=filearray[0]
fpre_lishi=filearray[1]
fpre_wuli=filearray[2]
fcur_lishi=filearray[3]
fcur_wuli=filearray[4]

wb_lishi = Workbook()
ws_lishi = wb_lishi.active
ws_lishi.title='Sheet1'
ws_lishi.append(biaotou)

wb_wuli = Workbook()
ws_wuli = wb_wuli.active
ws_wuli.title='Sheet1'
ws_wuli.append(biaotou)

pre_ref_df=pd.read_excel(ref)
pre_lishi_df=pd.read_excel(fpre_lishi)
pre_wuli_df=pd.read_excel(fpre_wuli)

cur_lishi_df=pd.read_excel(fcur_lishi, converters = {u'院校代码':str,u'专业代码':str})

cur_wuli_df=pd.read_excel(fcur_wuli, converters = {u'院校代码':str,u'专业代码':str})
#cur_wuli.guess_types = True
#cur_wuli_sheet = cur_wuli['Sheet1']
#cur_wuli_colum = cur_wuli_sheet.columns

matrix = []
#a = 0

for i, row in cur_lishi_df.iterrows():   
	matrix.append(row['院校代码'])    #'院校代码'
	matrix.append(row['院校'])    #'院校名称'
	matrix.append(row['专业代码'])    #'专业代码'
	matrix.append(row['专业'])    #'专业名称'
	matrix.append(row['专业备注'])    #'专业名称'
	matrix.append(row['计划数'])      #'计划数'
	matrix.append(row['类型'])        #'类型'
	matrix.append(row['性质'])        #'性质'
	matrix.append(row['选科'])        #'选科'
	matrix.append(row['院校省份'])    #'省份'	
	matrix.append(row['院校城市'])    #'城市'
	matrix.append(row['学费'])        #'学费'
	
	pre_score = pre_lishi_df.loc[ (pre_lishi_df['院校名称'].str.contains(matrix[1])) & (pre_lishi_df['专业名称'] == matrix[3]) , ['投档分'] ]
	
	if not pre_score.empty: 
		matrix.append(pre_score.iloc[0, 0])   #'去年提档分数'
		pre_rank = pre_ref_df.loc[ pre_ref_df['分数'] == pre_score.iloc[0, 0], ['历史位次']]
		if not pre_rank.empty: matrix.append(pre_rank.iloc[0, 0])   #'去年提档位次'
	
	#a += 1
	ws_lishi.append(matrix)     
	matrix = []
	#if a == 20:
		#break
  
wb_lishi.save('E:\\Tools\\GaoKao_files\\2023_pool_Lishi.xlsx')

matrix = []
#a = 0

for i, row in cur_wuli_df.iterrows():   
	matrix.append(row['院校代码'])    #'院校代码'
	matrix.append(row['院校'])    #'院校名称'
	matrix.append(row['专业代码'])    #'专业代码'
	matrix.append(row['专业'])    #'专业名称'
	matrix.append(row['专业备注'])    #'专业名称'
	matrix.append(row['计划数'])      #'计划数'
	matrix.append(row['类型'])        #'类型'
	matrix.append(row['性质'])        #'性质'
	matrix.append(row['选科'])        #'选科'
	matrix.append(row['院校省份'])    #'省份'	
	matrix.append(row['院校城市'])    #'城市'
	matrix.append(row['学费'])        #'学费'
	
	pre_score = pre_wuli_df.loc[ (pre_wuli_df['院校名称'].str.contains(matrix[1])) & (pre_wuli_df['专业名称'] == matrix[3]) , ['投档分'] ]
	
	if not pre_score.empty: 
		matrix.append(pre_score.iloc[0, 0])   #'去年提档分数'
		pre_rank = pre_ref_df.loc[ pre_ref_df['分数'] == pre_score.iloc[0, 0], ['物理位次']]
		if not pre_rank.empty: matrix.append(pre_rank.iloc[0, 0])   #'去年提档位次'
	
	#a += 1
	ws_wuli.append(matrix)     
	matrix = []
	#if a == 20:
		#break
  
wb_wuli.save('E:\\Tools\\GaoKao_files\\2023_pool_Wuli.xlsx')
